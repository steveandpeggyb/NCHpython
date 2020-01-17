from openpyxl.workbook import workbook
from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\csb003\\Documents\\NCHpython\\OpenPyXL\\regions.xlsx')
ws = wb.active

cell_range = ws['A1':'C1']
col_c = ws['C']
col_range = ws['A' : 'C']
row_range = ws[1:5]

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    for cell in row:
        print(cell)


